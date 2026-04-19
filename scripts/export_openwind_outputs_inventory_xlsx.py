from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "docs" / "openwind_outputs_inventory.xlsx"


ROWS = [
    {
        "Categoria": "Outputs base directos",
        "Nombre": "result.frequencies",
        "Tipo de miembro": "atributo",
        "Tipo observado": "ndarray",
        "Contenido / comportamiento": "Eje de frecuencias usado en el calculo.",
        "Serializacion": "Facil",
        "Costo adicional": "Bajo o nulo",
        "Recomendacion": "Guardar",
        "Notas": "Output base del calculo.",
    },
    {
        "Categoria": "Outputs base directos",
        "Nombre": "result.impedance",
        "Tipo de miembro": "atributo",
        "Tipo observado": "ndarray complejo",
        "Contenido / comportamiento": "Impedancia compleja de entrada.",
        "Serializacion": "Posible separando parte real e imaginaria",
        "Costo adicional": "Bajo o nulo",
        "Recomendacion": "Guardar",
        "Notas": "Output base del calculo; valioso para replicacion.",
    },
    {
        "Categoria": "Outputs base directos",
        "Nombre": "result.Zc",
        "Tipo de miembro": "atributo",
        "Tipo observado": "float64",
        "Contenido / comportamiento": "Impedancia caracteristica real.",
        "Serializacion": "Trivial",
        "Costo adicional": "Bajo o nulo",
        "Recomendacion": "Guardar",
        "Notas": "Output base del calculo.",
    },
    {
        "Categoria": "Resonancias y antiresonancias",
        "Nombre": "resonance_frequencies(k=5)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "ndarray",
        "Contenido / comportamiento": "Lista de frecuencias resonantes.",
        "Serializacion": "Facil",
        "Costo adicional": "Moderado",
        "Recomendacion": "Guardar",
        "Notas": "Metodo publico actual principal.",
    },
    {
        "Categoria": "Resonancias y antiresonancias",
        "Nombre": "resonance_peaks(k=5)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "tuple(ndarray, ndarray, ndarray)",
        "Contenido / comportamiento": "Frecuencias, factores y amplitudes/picos.",
        "Serializacion": "Facil separando arrays",
        "Costo adicional": "Moderado",
        "Recomendacion": "Guardar opcionalmente",
        "Notas": "Puede ser util para analisis mas fino.",
    },
    {
        "Categoria": "Resonancias y antiresonancias",
        "Nombre": "antiresonance_frequencies(k=5)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "ndarray",
        "Contenido / comportamiento": "Lista de antiresonancias.",
        "Serializacion": "Facil",
        "Costo adicional": "Moderado",
        "Recomendacion": "Opcional",
        "Notas": "No esencial para la primera version.",
    },
    {
        "Categoria": "Resonancias y antiresonancias",
        "Nombre": "antiresonance_peaks(k=5)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "tuple(ndarray, ndarray, list)",
        "Contenido / comportamiento": "Antiresonancias y picos asociados.",
        "Serializacion": "Factible",
        "Costo adicional": "Moderado",
        "Recomendacion": "Opcional",
        "Notas": "Valioso solo si se necesita analisis extendido.",
    },
    {
        "Categoria": "Informacion tecnica del solver",
        "Nombre": "technical_infos()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "None",
        "Contenido / comportamiento": "Imprime informacion tecnica a consola.",
        "Serializacion": "No directa",
        "Costo adicional": "Bajo",
        "Recomendacion": "No guardar directamente",
        "Notas": "Util para lectura humana, no como output estructurado.",
    },
    {
        "Categoria": "Informacion tecnica del solver",
        "Nombre": "discretization_infos()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "None",
        "Contenido / comportamiento": "Imprime informacion de discretizacion a consola.",
        "Serializacion": "No directa",
        "Costo adicional": "Bajo",
        "Recomendacion": "No guardar directamente",
        "Notas": "Mejor reconstruir o capturar aparte si hace falta.",
    },
    {
        "Categoria": "Informacion tecnica del solver",
        "Nombre": "get_nb_dof()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "int64",
        "Contenido / comportamiento": "Numero de grados de libertad.",
        "Serializacion": "Trivial",
        "Costo adicional": "Bajo",
        "Recomendacion": "Guardar opcionalmente",
        "Notas": "Bueno para auditoria tecnica.",
    },
    {
        "Categoria": "Informacion tecnica del solver",
        "Nombre": "get_entry_coefs(*labels)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "tuple",
        "Contenido / comportamiento": "Coeficientes de entrada; en la prueba salio vacio.",
        "Serializacion": "Posible",
        "Costo adicional": "Incierto",
        "Recomendacion": "No priorizar",
        "Notas": "No es clave en el flujo actual.",
    },
    {
        "Categoria": "Geometria y etiquetas",
        "Nombre": "get_instrument_geometry()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "InstrumentGeometry",
        "Contenido / comportamiento": "Objeto de geometria interna.",
        "Serializacion": "No recomendable como objeto",
        "Costo adicional": "Bajo",
        "Recomendacion": "No guardar el objeto",
        "Notas": "Ya guardamos main_bore y holes_valves.",
    },
    {
        "Categoria": "Geometria y etiquetas",
        "Nombre": "get_all_notes()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "list",
        "Contenido / comportamiento": "Notas del chart; vacio en nuestra prueba.",
        "Serializacion": "Facil",
        "Costo adicional": "Bajo",
        "Recomendacion": "No priorizar",
        "Notas": "Bajo valor sin digitacion real.",
    },
    {
        "Categoria": "Geometria y etiquetas",
        "Nombre": "get_pipes_label()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "list",
        "Contenido / comportamiento": "Etiquetas de pipes.",
        "Serializacion": "Facil",
        "Costo adicional": "Bajo",
        "Recomendacion": "Guardar opcionalmente",
        "Notas": "Util para auditoria tecnica.",
    },
    {
        "Categoria": "Geometria y etiquetas",
        "Nombre": "get_connectors_label()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "list",
        "Contenido / comportamiento": "Etiquetas de conectores.",
        "Serializacion": "Facil",
        "Costo adicional": "Bajo",
        "Recomendacion": "Guardar opcionalmente",
        "Notas": "Util para diagnostico tecnico.",
    },
    {
        "Categoria": "Geometria y etiquetas",
        "Nombre": "get_components_label()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "list",
        "Contenido / comportamiento": "Etiquetas de componentes.",
        "Serializacion": "Facil",
        "Costo adicional": "Bajo",
        "Recomendacion": "Guardar opcionalmente",
        "Notas": "Util para diagnostico tecnico.",
    },
    {
        "Categoria": "Campos acusticos",
        "Nombre": "get_pressure_flow()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "tuple(ndarray, ndarray, ndarray) en FEM+interp",
        "Contenido / comportamiento": "Ubicacion espacial, presion compleja y flujo complejo.",
        "Serializacion": "Pesada pero posible",
        "Costo adicional": "Alto",
        "Recomendacion": "No guardar por defecto",
        "Notas": "En modal no disponible por falta de interpolacion.",
    },
    {
        "Categoria": "Campos acusticos",
        "Nombre": "get_energy_field()",
        "Tipo de miembro": "metodo",
        "Tipo observado": "tuple(ndarray, ndarray) en FEM+interp",
        "Contenido / comportamiento": "Campo de energia espacial.",
        "Serializacion": "Pesada pero posible",
        "Costo adicional": "Alto",
        "Recomendacion": "No guardar por defecto",
        "Notas": "En modal no disponible en la practica actual.",
    },
    {
        "Categoria": "Evaluacion adicional",
        "Nombre": "evaluate_impedance_at(freqs)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "ndarray complejo en modal",
        "Contenido / comportamiento": "Evalua impedancia en frecuencias puntuales nuevas.",
        "Serializacion": "Posible",
        "Costo adicional": "Moderado",
        "Recomendacion": "No guardar como principal",
        "Notas": "Mejor usarlo como herramienta de analisis/refinamiento.",
    },
    {
        "Categoria": "Visualizacion / escritura",
        "Nombre": "plot_impedance(), plot_admittance(), plot_ac_field(), plot_ac_field_at_freq(), plot_instrument_geometry()",
        "Tipo de miembro": "metodos",
        "Tipo observado": "visualizacion",
        "Contenido / comportamiento": "Generan figuras/graficos.",
        "Serializacion": "No aplica como salida principal",
        "Costo adicional": "Variable",
        "Recomendacion": "No guardar en base",
        "Notas": "Solo para inspeccion visual.",
    },
    {
        "Categoria": "Visualizacion / escritura",
        "Nombre": "write_impedance(...)",
        "Tipo de miembro": "metodo",
        "Tipo observado": "escritura a archivo",
        "Contenido / comportamiento": "Exporta impedancia a un archivo.",
        "Serializacion": "Ya es exportacion",
        "Costo adicional": "Moderado",
        "Recomendacion": "No usar como base principal",
        "Notas": "Mejor guardar estructurado en DB.",
    },
]


SUMMARY = {
    "Guardar siempre": [
        "frequencies",
        "impedance",
        "Zc",
        "resonance_frequencies",
        "f1",
        "f2",
        "delta_cents",
    ],
    "Guardar opcionalmente": [
        "resonance_peaks",
        "antiresonance_frequencies",
        "antiresonance_peaks",
        "get_nb_dof",
        "get_pipes_label / get_connectors_label / get_components_label",
    ],
    "No guardar por defecto": [
        "technical_infos()",
        "discretization_infos()",
        "InstrumentGeometry como objeto",
        "pressure / flow / energy",
        "plotting methods",
    ],
}

SELECTED_OUTPUTS = [
    ("frequencies", "Si", "Output base directo de la clase."),
    ("impedance", "Si", "Output base directo de la clase; se guarda serializado por corte."),
    ("Zc", "Si", "Output base directo de la clase."),
    ("f1", "Si", "Derivado del pipeline usando resonance_frequencies()."),
    ("f2", "Si", "Derivado del pipeline usando resonance_frequencies()."),
    ("delta_cents", "Si", "Inarmonicidad derivada de f1 y f2."),
    ("resonance_peaks", "No por ahora", "Valioso pero no imprescindible en esta etapa."),
    ("antiresonance_frequencies", "No por ahora", "Opcional."),
    ("pressure_flow", "No por defecto", "Pesado y no disponible en modal."),
    ("energy_field", "No por defecto", "Pesado y no disponible en modal."),
]


def style_header(row):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def autofit(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = list(ROWS[0].keys())
    ws.append(headers)
    style_header(ws[1])

    for row in ROWS:
        ws.append([row[h] for h in headers])

    ws.freeze_panes = "A2"
    autofit(ws)

    summary_ws = wb.create_sheet("Resumen")
    summary_ws.append(["Grupo", "Elemento"])
    style_header(summary_ws[1])
    for group, items in SUMMARY.items():
        for item in items:
            summary_ws.append([group, item])
    summary_ws.freeze_panes = "A2"
    autofit(summary_ws)

    selected_ws = wb.create_sheet("Seleccionados")
    selected_ws.append(["Output", "Va a la base", "Motivo"])
    style_header(selected_ws[1])
    for row in SELECTED_OUTPUTS:
        selected_ws.append(list(row))
    selected_ws.freeze_panes = "A2"
    autofit(selected_ws)

    wb.save(OUTPUT_PATH)
    print(f"Excel exported to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
